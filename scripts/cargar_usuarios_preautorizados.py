"""
cargar_usuarios_preautorizados.py — Carga la lista de correos preautorizados
(rol + unidad/entidad) desde los Excel de solicitud de datos ya llenados por
las entidades, para habilitar el autoservicio de alta (POST /auth/solicitar-acceso).

Archivos esperados (carpeta solicitudes_altas_usuarios/, no se versiona en git):
    - layout_responsables_unidad.xlsx       (hoja "Responsables de Unidad")
          columnas: nombre_completo, cedula_profesional, correo_institucional,
                     clues_unidad, nombre_unidad, cargo_puesto, telefono_contacto
          -> rol_nombre = RESPONSABLE_UNIDAD, clues_unidad_asignada = clues_unidad
    - layout_administradores_estatales.xlsx (hoja "Administradores Estatales")
          columnas: nombre_completo, correo_institucional, entidad_federativa,
                     cargo_puesto, telefono_contacto
          -> rol_nombre = ADMIN_ESTATAL, id_entidad = entidad_federativa normalizada
             (MAYÚSCULAS, sin acentos — misma convención que cat_unidades.id_entidad)

El script es idempotente: si el correo ya está preautorizado, actualiza
rol/CLUES/entidad en vez de duplicar. Si el correo ya tiene una cuenta
`Usuario` creada, lo reporta como aviso (la preautorización no tiene efecto
práctico en ese caso, pero no es un error).

Dependencias:
    pip install openpyxl

Uso:
    python scripts/cargar_usuarios_preautorizados.py
    python scripts/cargar_usuarios_preautorizados.py --responsables otra_ruta.xlsx --administradores otra_ruta.xlsx
"""
import argparse
import sys
import unicodedata
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))

import openpyxl
from sqlalchemy.orm import Session

from app.database import SessionLocal
from app.models import UnidadMedica, Usuario, UsuarioPreautorizado

DATA_DIR = Path(__file__).parent.parent / "solicitudes_altas_usuarios"
ARCHIVO_RESPONSABLES_DEFAULT = DATA_DIR / "layout_responsables_unidad.xlsx"
ARCHIVO_ADMINISTRADORES_DEFAULT = DATA_DIR / "layout_administradores_estatales.xlsx"

ROL_RESPONSABLE_UNIDAD = "RESPONSABLE_UNIDAD"
ROL_ADMIN_ESTATAL = "ADMIN_ESTATAL"


def _normalizar_texto(texto: str) -> str:
    """MAYÚSCULAS sin acentos, conservando espacios (misma convención que
    cat_unidades.id_entidad, ej. 'CIUDAD DE MEXICO')."""
    sin_acentos = unicodedata.normalize("NFD", texto.strip().upper())
    return "".join(c for c in sin_acentos if unicodedata.category(c) != "Mn")


def _leer_hoja(ruta_excel: Path, nombre_hoja: str, columnas_requeridas: list[str]):
    wb = openpyxl.load_workbook(ruta_excel, data_only=True, read_only=True)
    if nombre_hoja not in wb.sheetnames:
        wb.close()
        raise ValueError(f"El Excel '{ruta_excel.name}' no tiene la hoja '{nombre_hoja}'.")
    ws = wb[nombre_hoja]

    encabezados = [
        str(cell.value).strip().lower() if cell.value else ""
        for cell in next(ws.iter_rows(min_row=1, max_row=1))
    ]
    faltantes = [c for c in columnas_requeridas if c not in encabezados]
    if faltantes:
        wb.close()
        raise ValueError(
            f"El Excel '{ruta_excel.name}' no tiene las columnas {faltantes}. "
            f"Encabezados encontrados: {encabezados}"
        )

    indices = {col: encabezados.index(col) for col in columnas_requeridas}
    filas = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()
    return indices, filas


def cargar_responsables_unidad(db: Session, ruta_excel: Path) -> dict:
    resultado = {"insertados": 0, "actualizados": 0, "invalidos": 0, "ya_registrados": 0, "errores": []}

    if not ruta_excel.exists():
        print(f"  [--] No se encontró {ruta_excel} — se omite carga de Responsables de Unidad.")
        return resultado

    indices, filas = _leer_hoja(
        ruta_excel, "Responsables de Unidad",
        ["correo_institucional", "clues_unidad"],
    )
    clues_validas = {r[0] for r in db.query(UnidadMedica.clues).all()}
    usuarios_existentes = {r[0] for r in db.query(Usuario.email).all()}

    for fila_num, fila in enumerate(filas, start=2):
        email_val = fila[indices["correo_institucional"]]
        clues_val = fila[indices["clues_unidad"]]
        if email_val is None and clues_val is None:
            continue

        email = str(email_val).strip().lower() if email_val else ""
        clues = str(clues_val).strip().upper() if clues_val else ""

        if not email or not clues:
            resultado["errores"].append(f"  Fila {fila_num}: correo o CLUES vacío.")
            resultado["invalidos"] += 1
            continue

        if clues not in clues_validas:
            resultado["errores"].append(f"  Fila {fila_num}: CLUES '{clues}' no existe en cat_unidades.")
            resultado["invalidos"] += 1
            continue

        existente = db.query(UsuarioPreautorizado).filter(UsuarioPreautorizado.email == email).first()
        if existente:
            existente.rol_nombre = ROL_RESPONSABLE_UNIDAD
            existente.clues_unidad_asignada = clues
            existente.id_entidad = None
            resultado["actualizados"] += 1
        else:
            db.add(UsuarioPreautorizado(
                email=email,
                rol_nombre=ROL_RESPONSABLE_UNIDAD,
                clues_unidad_asignada=clues,
                id_entidad=None,
            ))
            resultado["insertados"] += 1

        if email in usuarios_existentes:
            resultado["ya_registrados"] += 1

    return resultado


def cargar_administradores_estatales(db: Session, ruta_excel: Path) -> dict:
    resultado = {"insertados": 0, "actualizados": 0, "invalidos": 0, "ya_registrados": 0, "errores": []}

    if not ruta_excel.exists():
        print(f"  [--] No se encontró {ruta_excel} — se omite carga de Administradores Estatales.")
        return resultado

    indices, filas = _leer_hoja(
        ruta_excel, "Administradores Estatales",
        ["correo_institucional", "entidad_federativa"],
    )
    entidades_validas = {r[0] for r in db.query(UnidadMedica.id_entidad).distinct().all()}
    usuarios_existentes = {r[0] for r in db.query(Usuario.email).all()}

    for fila_num, fila in enumerate(filas, start=2):
        email_val = fila[indices["correo_institucional"]]
        entidad_val = fila[indices["entidad_federativa"]]
        if email_val is None and entidad_val is None:
            continue

        email = str(email_val).strip().lower() if email_val else ""
        entidad = _normalizar_texto(str(entidad_val)) if entidad_val else ""

        if not email or not entidad:
            resultado["errores"].append(f"  Fila {fila_num}: correo o entidad vacío.")
            resultado["invalidos"] += 1
            continue

        if entidad not in entidades_validas:
            resultado["errores"].append(
                f"  Fila {fila_num}: entidad '{entidad}' no coincide con ninguna de cat_unidades.id_entidad."
            )
            resultado["invalidos"] += 1
            continue

        existente = db.query(UsuarioPreautorizado).filter(UsuarioPreautorizado.email == email).first()
        if existente:
            existente.rol_nombre = ROL_ADMIN_ESTATAL
            existente.clues_unidad_asignada = None
            existente.id_entidad = entidad
            resultado["actualizados"] += 1
        else:
            db.add(UsuarioPreautorizado(
                email=email,
                rol_nombre=ROL_ADMIN_ESTATAL,
                clues_unidad_asignada=None,
                id_entidad=entidad,
            ))
            resultado["insertados"] += 1

        if email in usuarios_existentes:
            resultado["ya_registrados"] += 1

    return resultado


def _imprimir_resultado(titulo: str, resultado: dict) -> None:
    print(f"\n  {titulo}")
    print(f"    Insertados     : {resultado['insertados']}")
    print(f"    Actualizados   : {resultado['actualizados']}")
    print(f"    Inválidos      : {resultado['invalidos']}")
    print(f"    Ya con cuenta  : {resultado['ya_registrados']} (preautorizados que ya tienen Usuario — sin efecto práctico)")
    if resultado["errores"]:
        print(f"    Detalle de filas inválidas (máximo 20):")
        for e in resultado["errores"][:20]:
            print(e)
        if len(resultado["errores"]) > 20:
            print(f"    ... y {len(resultado['errores']) - 20} filas más.")


def cargar(ruta_responsables: Path, ruta_administradores: Path) -> None:
    print("\n" + "=" * 60)
    print("  CARGA — USUARIOS PREAUTORIZADOS")
    print("=" * 60)

    db: Session = SessionLocal()
    try:
        resultado_resp = cargar_responsables_unidad(db, ruta_responsables)
        resultado_admin = cargar_administradores_estatales(db, ruta_administradores)
        db.commit()

        _imprimir_resultado("Responsables de Unidad", resultado_resp)
        _imprimir_resultado("Administradores Estatales", resultado_admin)

        print("\n  Carga completada.\n" + "=" * 60 + "\n")
    except Exception as ex:
        db.rollback()
        print(f"\n  [ERROR] {ex}")
        raise
    finally:
        db.close()


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Carga usuarios preautorizados desde Excel.")
    parser.add_argument("--responsables", type=Path, default=ARCHIVO_RESPONSABLES_DEFAULT)
    parser.add_argument("--administradores", type=Path, default=ARCHIVO_ADMINISTRADORES_DEFAULT)
    args = parser.parse_args()
    cargar(args.responsables, args.administradores)
