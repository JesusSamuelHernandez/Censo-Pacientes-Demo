"""
scripts/crear_plantillas_excel.py — Genera los archivos Excel de plantilla
para la carga masiva de unidades médicas y medicamentos.

Uso:
    python scripts/crear_plantillas_excel.py

Genera en scripts/data/:
    unidades.xlsx      — Plantilla con encabezados y 3 filas de ejemplo.
    medicamentos.xlsx  — Plantilla con encabezados y 3 filas de ejemplo.

Una vez generados, llena las filas con tus datos reales y ejecuta:
    python scripts/cargar_unidades.py
    python scripts/cargar_medicamentos.py
"""
import os

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Directorio de salida
DATA_DIR = os.path.join(os.path.dirname(__file__), "data")
os.makedirs(DATA_DIR, exist_ok=True)


# ---------------------------------------------------------------------------
# Helpers de estilo
# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
EXAMPLE_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
EXAMPLE_FONT = Font(color="1F4E79", italic=True, size=10)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _estilizar_encabezado(ws, headers: list[str], col_widths: list[int]) -> None:
    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
        ws.column_dimensions[cell.column_letter].width = width
    ws.row_dimensions[1].height = 22


def _estilizar_fila_ejemplo(ws, row_idx: int, num_cols: int) -> None:
    for col_idx in range(1, num_cols + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.font = EXAMPLE_FONT
        cell.fill = EXAMPLE_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical="center")


# ---------------------------------------------------------------------------
# Plantilla: unidades.xlsx
# ---------------------------------------------------------------------------
def crear_plantilla_unidades() -> str:
    ruta = os.path.join(DATA_DIR, "unidades.xlsx")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Unidades"

    headers = ["clues", "nombre_de_la_unidad", "id_entidad", "categoria_gerencial"]
    widths = [18, 55, 25, 22]
    _estilizar_encabezado(ws, headers, widths)

    ejemplos = [
        ["BCSSA004266", "Hospital General Tijuana",                  "BAJA_CALIFORNIA",    "HG"],
        ["JCSSA000180", "Centro de Salud Guadalajara Norte",         "JALISCO",            "CS"],
        ["MCSSA002451", "Hospital Comunitario Morelia",              "MICHOACAN",          "HC"],
        ["DFSSA001234", "Unidad de Especialidades Médicas CDMX",    "CIUDAD_DE_MEXICO",   "UNEME"],
        ["NLSSA003789", "Hospital General Monterrey",                "NUEVO_LEON",         "HG"],
    ]

    for row_idx, fila in enumerate(ejemplos, start=2):
        for col_idx, valor in enumerate(fila, start=1):
            ws.cell(row=row_idx, column=col_idx, value=valor)
        _estilizar_fila_ejemplo(ws, row_idx, len(headers))

    ws.freeze_panes = "A2"
    wb.save(ruta)
    return ruta


# ---------------------------------------------------------------------------
# Plantilla: medicamentos.xlsx
# ---------------------------------------------------------------------------
def crear_plantilla_medicamentos() -> str:
    ruta = os.path.join(DATA_DIR, "medicamentos.xlsx")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Medicamentos"

    headers = ["clave_cnis", "descripcion", "grupo", "tipo_clave"]
    widths = [22, 65, 22, 25]
    _estilizar_encabezado(ws, headers, widths)

    ejemplos = [
        ["010.000.4155.00", "RITUXIMAB 500MG SOLUCION INYECTABLE",                    "BIOLOGICOS",   "CUADRO BASICO"],
        ["010.000.5230.00", "ADALIMUMAB 40MG/0.8ML SOLUCION INYECTABLE",              "BIOLOGICOS",   "CUADRO BASICO"],
        ["010.000.4892.00", "IMATINIB 400MG TABLETAS",                                "ONCOLOGICOS",  "GASTOS CATASTROFICOS"],
        ["010.000.5341.00", "TRASTUZUMAB 440MG SOLUCION INYECTABLE",                  "ONCOLOGICOS",  "GASTOS CATASTROFICOS"],
        ["010.000.4601.00", "ETANERCEPT 50MG SOLUCION INYECTABLE",                    "BIOLOGICOS",   "CUADRO BASICO"],
        ["010.000.5102.00", "BEVACIZUMAB 400MG/16ML SOLUCION INYECTABLE",             "ONCOLOGICOS",  "GASTOS CATASTROFICOS"],
        ["010.000.4780.00", "INFLIXIMAB 100MG POLVO LIOFILIZADO INYECTABLE",          "BIOLOGICOS",   "CUADRO BASICO"],
    ]

    for row_idx, fila in enumerate(ejemplos, start=2):
        for col_idx, valor in enumerate(fila, start=1):
            ws.cell(row=row_idx, column=col_idx, value=valor)
        _estilizar_fila_ejemplo(ws, row_idx, len(headers))

    ws.freeze_panes = "A2"
    wb.save(ruta)
    return ruta


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    print("\nGenerando plantillas Excel...\n")

    ruta_u = crear_plantilla_unidades()
    print(f"  [OK] {ruta_u}")

    ruta_m = crear_plantilla_medicamentos()
    print(f"  [OK] {ruta_m}")

    print("\nPlantillas listas. Instrucciones:")
    print("  1. Abre los archivos en Excel y reemplaza las filas de ejemplo con tus datos reales.")
    print("  2. Ejecuta la carga:")
    print("       python scripts/cargar_unidades.py")
    print("       python scripts/cargar_medicamentos.py")
    print("  3. Puedes pasar un archivo diferente con --archivo:")
    print("       python scripts/cargar_unidades.py --archivo C:/ruta/mis_unidades.xlsx\n")
