"""
cargar_medicos.py — Carga médicos desde un Excel y los registra en la BD.

El archivo esperado (carpeta solicitudes_altas_usuarios/, no se versiona en git):
    layout_medicos.xlsx   (hoja "Médicos")
        columnas obligatorias : nombre_completo, cedula_profesional, clues_unidad
        columnas opcionales   : curp, email, codigo_puesto

El script es idempotente: si ya existe un médico con la misma cédula (por
cedula_hash), actualiza sus datos en lugar de duplicarlo.

Los campos nombre_completo, cedula_profesional y curp se almacenan cifrados
(Fernet). La búsqueda de duplicados se hace por hash SHA-256 de la cédula.

Dependencias:
    pip install openpyxl

Uso:
    python scripts/cargar_medicos.py
    python scripts/cargar_medicos.py --archivo otra_ruta.xlsx
    python scripts/cargar_medicos.py --generar-plantilla
"""
import argparse
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))

# dotenv debe cargarse antes de importar cualquier módulo de app/
from dotenv import load_dotenv
load_dotenv(Path(__file__).parent.parent / ".env")

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from sqlalchemy.orm import Session

from app.crypto import cifrar, hash_sha256
from app.database import SessionLocal
from app.models import CatPuesto, Medico, UnidadMedica

DATA_DIR = Path(__file__).parent.parent / "solicitudes_altas_usuarios"
ARCHIVO_DEFAULT = DATA_DIR / "layout_medicos.xlsx"
NOMBRE_HOJA = "Médicos"

COLUMNAS_REQUERIDAS = ["nombre_completo", "cedula_profesional", "clues_unidad"]
COLUMNAS_OPCIONALES = ["curp", "email", "codigo_puesto"]
TODAS_COLUMNAS = COLUMNAS_REQUERIDAS + COLUMNAS_OPCIONALES

# ---------------------------------------------------------------------------
# Generador de plantilla Excel
# ---------------------------------------------------------------------------

HEADER_FILL  = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
EXAMPLE_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
REQ_FILL     = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
HEADER_FONT  = Font(color="FFFFFF", bold=True, size=11)
EXAMPLE_FONT = Font(color="1F4E79", italic=True, size=10)
TITLE_FONT   = Font(color="1F4E79", bold=True, size=13)
THIN_BORDER  = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)
COL_WIDTHS = [35, 22, 18, 22, 35, 18]


def generar_plantilla(ruta: Path) -> None:
    ruta.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()

    # ── Hoja de instrucciones ──────────────────────────────────────────────
    ws_inst = wb.active
    ws_inst.title = "Instrucciones"
    ws_inst.column_dimensions["A"].width = 90

    instrucciones = [
        ("LAYOUT DE CARGA MASIVA — MÉDICOS", TITLE_FONT),
        ("", None),
        ("Instrucciones de llenado:", Font(bold=True, size=11)),
        ("1. Llene la hoja 'Médicos' con los datos de cada médico, una fila por médico.", Font(size=11)),
        ("2. No modifique los encabezados de la hoja de datos.", Font(size=11)),
        ("3. Las columnas en AMARILLO son obligatorias; las demás son opcionales.", Font(size=11)),
        ("4. La cédula profesional debe ser el número oficial de la SEP (sin espacios).", Font(size=11)),
        ("5. La CLUES debe coincidir exactamente con el catálogo de unidades del sistema.", Font(size=11)),
        ("6. El código de puesto debe coincidir con el catálogo de puestos del sistema.", Font(size=11)),
        ("   (Consulte al administrador del sistema para obtener los códigos válidos.)", Font(size=11, italic=True)),
        ("7. La fila de ejemplo (en azul claro) puede eliminarse antes de entregar el archivo.", Font(size=11)),
        ("", None),
        ("Columnas obligatorias:", Font(bold=True, size=11)),
        ("  nombre_completo     — Nombre completo del médico (apellidos + nombre)", Font(size=11)),
        ("  cedula_profesional  — Número de cédula profesional SEP", Font(size=11)),
        ("  clues_unidad        — CLUES de la unidad médica de adscripción", Font(size=11)),
        ("", None),
        ("Columnas opcionales:", Font(bold=True, size=11)),
        ("  curp         — CURP del médico (18 caracteres)", Font(size=11)),
        ("  email        — Correo electrónico de contacto", Font(size=11)),
        ("  codigo_puesto — Código del puesto en el catálogo del sistema", Font(size=11)),
    ]

    for row_idx, (texto, fuente) in enumerate(instrucciones, start=1):
        cell = ws_inst.cell(row=row_idx, column=1, value=texto)
        if fuente:
            cell.font = fuente
        cell.alignment = Alignment(wrap_text=True)

    # ── Hoja de datos ──────────────────────────────────────────────────────
    ws = wb.create_sheet(title=NOMBRE_HOJA)

    encabezados_display = [
        "nombre_completo *",
        "cedula_profesional *",
        "clues_unidad *",
        "curp",
        "email",
        "codigo_puesto",
    ]

    for col_idx, (header, width) in enumerate(zip(encabezados_display, COL_WIDTHS), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL if not header.endswith("*") or True else REQ_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
        ws.column_dimensions[cell.column_letter].width = width
    ws.row_dimensions[1].height = 30

    ejemplo = [
        "LOPEZ HERNANDEZ MARIA ELENA",
        "12345678",
        "HGO012345678901",
        "LOHM850101MDFPRR09",
        "maria.lopez@imssbienestar.gob.mx",
        "M-ESP-01",
    ]

    for col_idx, valor in enumerate(ejemplo, start=1):
        cell = ws.cell(row=2, column=col_idx, value=valor)
        cell.font = EXAMPLE_FONT
        cell.fill = EXAMPLE_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[2].height = 20

    ws.freeze_panes = "A2"

    wb.save(ruta)
    print(f"Plantilla generada en: {ruta}")


# ---------------------------------------------------------------------------
# Carga desde Excel
# ---------------------------------------------------------------------------

def _leer_hoja(ruta_excel: Path):
    wb = openpyxl.load_workbook(ruta_excel, data_only=True, read_only=True)

    nombre_hoja_real = next(
        (h for h in wb.sheetnames if h.strip().lower() == NOMBRE_HOJA.lower()),
        None,
    )
    if not nombre_hoja_real:
        wb.close()
        raise ValueError(
            f"El Excel '{ruta_excel.name}' no tiene la hoja '{NOMBRE_HOJA}'. "
            f"Hojas encontradas: {wb.sheetnames}"
        )

    ws = wb[nombre_hoja_real]
    encabezados_raw = [
        str(cell.value).strip().lower().replace(" *", "").replace("*", "")
        if cell.value else ""
        for cell in next(ws.iter_rows(min_row=1, max_row=1))
    ]

    faltantes = [c for c in COLUMNAS_REQUERIDAS if c not in encabezados_raw]
    if faltantes:
        wb.close()
        raise ValueError(
            f"El Excel '{ruta_excel.name}' no tiene las columnas requeridas: {faltantes}. "
            f"Encabezados encontrados: {encabezados_raw}"
        )

    indices = {
        col: encabezados_raw.index(col)
        for col in TODAS_COLUMNAS
        if col in encabezados_raw
    }

    filas = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()
    return indices, filas


def cargar(ruta_excel: Path) -> None:
    print("\n" + "=" * 60)
    print("  CARGA — MEDICOS")
    print("=" * 60)

    if not ruta_excel.exists():
        print(f"\n  [--] No se encontro el archivo: {ruta_excel}")
        print(f"       Genera la plantilla con:")
        print(f"       python scripts/cargar_medicos.py --generar-plantilla")
        sys.exit(1)

    try:
        indices, filas = _leer_hoja(ruta_excel)
    except ValueError as e:
        print(f"\n  [ERROR] {e}")
        sys.exit(1)

    db: Session = SessionLocal()
    try:
        clues_validas  = {r[0] for r in db.query(UnidadMedica.clues).all()}
        puestos_validos = {r[0] for r in db.query(CatPuesto.codigo).all()}

        insertados  = 0
        actualizados = 0
        invalidos   = 0
        errores     = []

        for fila_num, fila in enumerate(filas, start=2):
            def _val(col):
                idx = indices.get(col)
                if idx is None or idx >= len(fila):
                    return None
                v = fila[idx]
                return str(v).strip() if v is not None else None

            nombre   = _val("nombre_completo")
            cedula   = _val("cedula_profesional")
            clues    = _val("clues_unidad")
            curp     = _val("curp")
            email    = _val("email")
            puesto   = _val("codigo_puesto")

            # Saltar filas vacías
            if not nombre and not cedula and not clues:
                continue

            # Validaciones obligatorias
            if not nombre:
                errores.append(f"  Fila {fila_num}: nombre_completo vacio.")
                invalidos += 1
                continue
            if not cedula:
                errores.append(f"  Fila {fila_num}: cedula_profesional vacia.")
                invalidos += 1
                continue
            if not clues:
                errores.append(f"  Fila {fila_num}: clues_unidad vacia.")
                invalidos += 1
                continue

            clues_upper = clues.upper()
            if clues_upper not in clues_validas:
                errores.append(f"  Fila {fila_num}: CLUES '{clues_upper}' no existe en cat_unidades.")
                invalidos += 1
                continue

            if puesto:
                if puesto not in puestos_validos:
                    errores.append(f"  Fila {fila_num}: codigo_puesto '{puesto}' no existe en cat_puestos.")
                    invalidos += 1
                    continue

            cedula_hash = hash_sha256(cedula)
            curp_upper  = curp.upper() if curp else None
            curp_hash   = hash_sha256(curp_upper) if curp_upper else None

            existente = db.query(Medico).filter(Medico.cedula_hash == cedula_hash).first()

            if existente:
                existente.nombre_medico       = nombre.upper()
                existente.cedula              = cedula
                existente.clues_adscripcion   = clues_upper
                existente.email               = email.lower() if email else existente.email
                existente.id_puesto           = puesto or existente.id_puesto
                if curp_upper:
                    existente.curp      = curp_upper
                    existente.curp_hash = curp_hash
                existente.es_activo = True
                actualizados += 1
            else:
                db.add(Medico(
                    cedula_hash         = cedula_hash,
                    nombre_medico       = nombre.upper(),
                    cedula              = cedula,
                    clues_adscripcion   = clues_upper,
                    email               = email.lower() if email else None,
                    id_puesto           = puesto or None,
                    curp                = curp_upper,
                    curp_hash           = curp_hash,
                    es_activo           = True,
                ))
                insertados += 1

        db.commit()

        print(f"\n  Insertados   : {insertados}")
        print(f"  Actualizados : {actualizados}")
        print(f"  Invalidos    : {invalidos}")

        if errores:
            print(f"\n  Detalle de filas invalidas (maximo 30):")
            for e in errores[:30]:
                print(e)
            if len(errores) > 30:
                print(f"  ... y {len(errores) - 30} filas mas.")

        print("\n  Carga completada.\n" + "=" * 60 + "\n")

    except Exception as ex:
        db.rollback()
        print(f"\n  [ERROR] {ex}")
        raise
    finally:
        db.close()


# ---------------------------------------------------------------------------
# Punto de entrada
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Carga medicos desde Excel.")
    parser.add_argument(
        "--archivo",
        type=Path,
        default=ARCHIVO_DEFAULT,
        help=f"Ruta al Excel (default: {ARCHIVO_DEFAULT})",
    )
    parser.add_argument(
        "--generar-plantilla",
        action="store_true",
        help="Genera el layout Excel de ejemplo y termina.",
    )
    args = parser.parse_args()

    if args.generar_plantilla:
        generar_plantilla(ARCHIVO_DEFAULT)
    else:
        cargar(args.archivo)
